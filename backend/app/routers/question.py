from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
import json
import io

from app.core.database import get_db
from app.schemas.schemas import QuestionCreate, QuestionUpdate
from app.crud import crud
from app.models.models import Question, Category, Tag

router = APIRouter(prefix="/questions", tags=["题库管理"])


@router.get("")
def list_questions(
    page: int = 1,
    page_size: int = 20,
    question_type: Optional[str] = None,
    category_id: Optional[int] = None,
    difficulty: Optional[str] = None,
    tag_ids: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
):
    tag_id_list = []
    if tag_ids:
        try:
            tag_id_list = [int(x) for x in tag_ids.split(",") if x]
        except ValueError:
            tag_id_list = []

    skip = (page - 1) * page_size
    total, questions = crud.get_questions(
        db,
        skip=skip,
        limit=page_size,
        question_type=question_type,
        category_id=category_id,
        difficulty=difficulty,
        tag_ids=tag_id_list if tag_id_list else None,
        keyword=keyword,
    )

    items = []
    for q in questions:
        items.append({
            "id": q.id,
            "question_type": q.question_type,
            "title": q.title,
            "options": q.options,
            "answer": q.answer,
            "analysis": q.analysis,
            "difficulty": q.difficulty,
            "category_id": q.category_id,
            "category": {"id": q.category.id, "name": q.category.name} if q.category else None,
            "tags": [{"id": t.id, "name": t.name, "color": t.color} for t in q.tags],
            "created_at": q.created_at.isoformat(),
            "updated_at": q.updated_at.isoformat(),
        })

    return {
        "code": 200,
        "data": {
            "total": total,
            "items": items,
            "page": page,
            "page_size": page_size,
        },
        "msg": "success",
    }


@router.get("/{question_id}")
def get_question(question_id: int, db: Session = Depends(get_db)):
    q = crud.get_question(db, question_id)
    if not q:
        return {"code": 404, "msg": "题目不存在"}
    return {
        "code": 200,
        "data": {
            "id": q.id,
            "question_type": q.question_type,
            "title": q.title,
            "options": q.options,
            "answer": q.answer,
            "analysis": q.analysis,
            "difficulty": q.difficulty,
            "category_id": q.category_id,
            "category": {"id": q.category.id, "name": q.category.name} if q.category else None,
            "tags": [{"id": t.id, "name": t.name, "color": t.color} for t in q.tags],
            "created_at": q.created_at.isoformat(),
            "updated_at": q.updated_at.isoformat(),
        },
        "msg": "success",
    }


@router.post("")
def create_question(question: QuestionCreate, db: Session = Depends(get_db)):
    q = crud.create_question(db, question)
    return {
        "code": 200,
        "data": {"id": q.id},
        "msg": "创建成功",
    }


@router.put("/{question_id}")
def update_question(question_id: int, question: QuestionUpdate, db: Session = Depends(get_db)):
    q = crud.update_question(db, question_id, question)
    if not q:
        return {"code": 404, "msg": "题目不存在"}
    return {"code": 200, "data": {"id": q.id}, "msg": "更新成功"}


@router.delete("/{question_id}")
def delete_question(question_id: int, db: Session = Depends(get_db)):
    success = crud.delete_question(db, question_id)
    if not success:
        return {"code": 404, "msg": "题目不存在"}
    return {"code": 200, "msg": "删除成功"}


@router.post("/batch-delete")
def batch_delete(ids: List[int], db: Session = Depends(get_db)):
    crud.batch_delete_questions(db, ids)
    return {"code": 200, "msg": "批量删除成功"}


@router.post("/{question_id}/copy")
def copy_question(question_id: int, db: Session = Depends(get_db)):
    q = crud.copy_question(db, question_id)
    if not q:
        return {"code": 404, "msg": "题目不存在"}
    return {"code": 200, "data": {"id": q.id}, "msg": "复制成功"}


@router.post("/import/json")
async def import_json(file: UploadFile = File(...), db: Session = Depends(get_db)):
    content = await file.read()
    try:
        data = json.loads(content)
        if isinstance(data, dict) and "questions" in data:
            questions_data = data["questions"]
        elif isinstance(data, list):
            questions_data = data
        else:
            return {"code": 400, "msg": "JSON格式不正确"}

        count = 0
        for q_data in questions_data:
            try:
                # 转换 answer 为数组
                if 'answer' in q_data and isinstance(q_data['answer'], str):
                    q_data['answer'] = [q_data['answer']]

                # 转换 tags 名称列表为 tag_ids
                if 'tags' in q_data and isinstance(q_data['tags'], list):
                    tag_ids = []
                    for tag_name in q_data['tags']:
                        # 查找或创建标签
                        existing_tag = db.query(Tag).filter(Tag.name == tag_name).first()
                        if existing_tag:
                            tag_ids.append(existing_tag.id)
                        else:
                            new_tag = Tag(name=tag_name, color="#3b82f6")
                            db.add(new_tag)
                            db.flush()
                            tag_ids.append(new_tag.id)
                    q_data['tag_ids'] = tag_ids
                    del q_data['tags']

                # 转换 category_name 为 category_id
                if 'category_name' in q_data:
                    cat_name = q_data.pop('category_name')
                    if cat_name:
                        existing_cat = db.query(Category).filter(Category.name == cat_name).first()
                        if existing_cat:
                            q_data['category_id'] = existing_cat.id
                        else:
                            new_cat = Category(name=cat_name)
                            db.add(new_cat)
                            db.flush()
                            q_data['category_id'] = new_cat.id

                question = QuestionCreate(**q_data)
                crud.create_question(db, question)
                count += 1
            except Exception as e:
                print(f"导入题目失败: {e}, 数据: {q_data}")
                continue

        return {"code": 200, "data": {"imported": count}, "msg": f"成功导入 {count} 道题目"}
    except json.JSONDecodeError:
        return {"code": 400, "msg": "JSON解析失败"}


@router.get("/export/json")
def export_json(
    db: Session = Depends(get_db),
    question_type: Optional[str] = None,
    category_id: Optional[int] = None,
    difficulty: Optional[str] = None,
):
    _, questions = crud.get_questions(
        db,
        skip=0,
        limit=10000,
        question_type=question_type,
        category_id=category_id,
        difficulty=difficulty,
    )

    result = []
    for q in questions:
        result.append({
            "question_type": q.question_type,
            "title": q.title,
            "options": q.options,
            "answer": q.answer,
            "analysis": q.analysis,
            "difficulty": q.difficulty,
            "category_name": q.category.name if q.category else None,
            "tags": [t.name for t in q.tags],
        })

    return {"code": 200, "data": result, "msg": "success"}


@router.get("/export/markdown")
def export_markdown(
    db: Session = Depends(get_db),
    question_type: Optional[str] = None,
    category_id: Optional[int] = None,
    difficulty: Optional[str] = None,
):
    _, questions = crud.get_questions(
        db,
        skip=0,
        limit=10000,
        question_type=question_type,
        category_id=category_id,
        difficulty=difficulty,
    )

    md_content = "# 题库导出\n\n"
    type_map = {
        "single": "单选题",
        "multiple": "多选题",
        "judge": "判断题",
        "short_answer": "简答题",
    }

    for i, q in enumerate(questions, 1):
        type_name = type_map.get(q.question_type, q.question_type)
        md_content += f"## {i}. [{type_name}] {q.title}\n\n"

        if q.options and q.question_type in ["single", "multiple"]:
            for idx, opt in enumerate(q.options):
                letter = chr(65 + idx)
                md_content += f"- {letter}. {opt}\n"
            md_content += "\n"

        ans = q.answer
        if isinstance(ans, list):
            ans = ", ".join(ans)
        md_content += f"**答案：** {ans}\n\n"

        if q.analysis:
            md_content += f"**解析：** {q.analysis}\n\n"

        md_content += "---\n\n"

    return {"code": 200, "data": {"content": md_content}, "msg": "success"}


@router.post("/import/xlsx")
async def import_xlsx(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """导入 xlsx 格式题库"""
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
    
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
        ws = wb.active
        
        # 获取表头（第一行）
        headers = [cell.value for cell in ws[1]]
        
        # 字段映射（支持多种表头格式）
        field_map = {
            '题型': 'question_type',
            '题目': 'title',
            '选项A': 'option_a',
            '选项B': 'option_b',
            '选项C': 'option_c',
            '选项D': 'option_d',
            '选项': 'options',  # 兼容旧格式
            '答案': 'answer',
            '解析': 'analysis',
            '难度': 'difficulty',
            '分类': 'category_name',
            '标签': 'tags',
        }
        
        # 将中文表头转换为英文字段
        header_to_field = {}
        for i, h in enumerate(headers):
            if h in field_map:
                header_to_field[i] = field_map[h]
        
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # 跳过空行
                continue
            
            q_data = {}
            options = []
            
            for col_idx, field in header_to_field.items():
                value = row[col_idx]
                if value is None:
                    continue
                
                if field == 'option_a':
                    options.append(str(value))
                elif field == 'option_b':
                    options.append(str(value))
                elif field == 'option_c':
                    options.append(str(value))
                elif field == 'option_d':
                    options.append(str(value))
                elif field == 'options':
                    # 旧格式：选项用换行符分隔
                    if isinstance(value, str):
                        opts = [opt.strip() for opt in value.split('\n') if opt.strip()]
                        options.extend(opts)
                elif field == 'tags':
                    # 标签用逗号分隔
                    if isinstance(value, str):
                        tags = [t.strip() for t in value.split(',') if t.strip()]
                        q_data[field] = tags
                else:
                    q_data[field] = str(value) if value else None
            
            # 合并选项
            if options:
                q_data['options'] = options
            
            if not q_data.get('title') or not q_data.get('question_type'):
                continue
            
            try:
                question = QuestionCreate(**q_data)
                crud.create_question(db, question)
                count += 1
            except Exception:
                continue
        
        return {"code": 200, "data": {"imported": count}, "msg": f"成功导入 {count} 道题目"}
    except InvalidFileException:
        return {"code": 400, "msg": "文件格式不正确，请使用 .xlsx 文件"}
    except Exception as e:
        return {"code": 400, "msg": f"导入失败: {str(e)}"}


@router.get("/export/xlsx")
def export_xlsx(
    db: Session = Depends(get_db),
    question_type: Optional[str] = None,
    category_id: Optional[int] = None,
    difficulty: Optional[str] = None,
):
    """导出 xlsx 格式题库"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    _, questions = crud.get_questions(
        db,
        skip=0,
        limit=10000,
        question_type=question_type,
        category_id=category_id,
        difficulty=difficulty,
    )
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题库"
    
    # 设置表头（不合并，每个选项独立列）
    headers = ['题型', '题目', '选项A', '选项B', '选项C', '选项D', '答案', '解析', '难度', '分类', '标签']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    for row_idx, q in enumerate(questions, 2):
        # 题型
        ws.cell(row=row_idx, column=1, value=q.question_type)
        
        # 题目
        ws.cell(row=row_idx, column=2, value=q.title)
        
        # 选项（每个选项独立一列）
        if q.options:
            for i, opt in enumerate(q.options[:4]):  # 最多4个选项
                ws.cell(row=row_idx, column=3 + i, value=opt)
        
        # 答案
        answer = q.answer
        if isinstance(answer, list):
            answer = ''.join(answer)
        ws.cell(row=row_idx, column=7, value=answer)
        
        # 解析
        ws.cell(row=row_idx, column=8, value=q.analysis or '')
        
        # 难度
        ws.cell(row=row_idx, column=9, value=q.difficulty or 'medium')
        
        # 分类
        ws.cell(row=row_idx, column=10, value=q.category.name if q.category else '')
        
        # 标签
        if q.tags:
            tags_text = ','.join([t.name for t in q.tags])
            ws.cell(row=row_idx, column=11, value=tags_text)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 20
    
    # 保存到BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename="questions.xlsx"'
        }
    )


@router.get("/template/xlsx")
def download_template():
    """下载 xlsx 导入模版"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题库导入模板"
    
    # 表头（不合并，每个选项独立列）
    headers = ['题型', '题目', '选项A', '选项B', '选项C', '选项D', '答案', '解析', '难度', '分类', '标签']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 填写示例数据（不合并单元格）
    example_data = [
        ['single', '以下哪个是正确的？', '选项1', '选项2', '选项3', '选项4', 'A', '解析：答案是A', 'medium', '示例分类', '标签1,标签2'],
        ['multiple', '以下哪些是正确的？', '选项1', '选项2', '选项3', '选项4', 'AB', '解析：答案是AB', 'easy', '示例分类', '标签3'],
        ['judge', '判断题：这是正确的', '', '', '', '', 'true', '解析：正确', 'hard', '', ''],
        ['short_answer', '简答题：请简述xxx', '', '', '', '', '答案内容', '评分要点：xxx', 'medium', '', ''],
    ]
    
    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 40
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 20
    
    # 说明行（第7行开始）
    ws.cell(row=7, column=1, value='--- 填写说明 ---')
    ws.cell(row=7, column=1).font = Font(bold=True)
    
    instructions = [
        '1. 题型：single(单选) / multiple(多选) / judge(判断) / short_answer(简答)',
        '2. 题目：题目内容，必填，支持 Markdown 格式',
        '3. 选项A-D：每个选项单独一列，单选/多选题必填',
        '4. 答案：单选如"A"，多选如"AB"，判断题填"true"或"false"',
        '5. 解析：答案解析，可为空',
        '6. 难度：easy / medium / hard，可为空默认为 medium',
        '7. 分类：分类名称，不存在则自动创建，可为空',
        '8. 标签：多个标签用逗号分隔，可为空',
        '9. 请删除示例数据后填写自己的题目',
    ]
    
    for i, text in enumerate(instructions, 8):
        ws.cell(row=i, column=1, value=text)
        ws.merge_cells(f'A{i}:K{i}')
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename="questions_template.xlsx"'
        }
    )
